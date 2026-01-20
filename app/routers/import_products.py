"""Import products from XLSX file"""
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Request
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from app.database import Book, Category, SessionLocal
from typing import List, Dict, Any
import io
import re
from app.admin.config import send_webhook

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def slugify(text: str) -> str:
    """Convert text to URL-friendly slug"""
    text = text.lower().strip()
    text = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', text)
    text = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', text)
    text = re.sub(r'[ìíịỉĩ]', 'i', text)
    text = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', text)
    text = re.sub(r'[ùúụủũưừứựửữ]', 'u', text)
    text = re.sub(r'[ỳýỵỷỹ]', 'y', text)
    text = re.sub(r'[đ]', 'd', text)
    text = re.sub(r'[^a-z0-9\s-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')

def get_or_create_category(db: Session, category_name: str) -> Category:
    """Get existing category or create new one"""
    if not category_name:
        return None
    
    category = db.query(Category).filter(Category.name == category_name).first()
    if not category:
        slug = slugify(category_name)
        # Make sure slug is unique
        existing = db.query(Category).filter(Category.slug == slug).first()
        if existing:
            slug = f"{slug}-{db.query(Category).count() + 1}"
        
        category = Category(
            name=category_name,
            slug=slug,
            description=f"Danh mục {category_name}"
        )
        db.add(category)
        db.commit()
        db.refresh(category)
    
    return category

@router.post("/import-products")
async def import_products_xlsx(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Import products from XLSX file"""
    # Check if user is admin (via session)
    session = request.session
    if not session.get("admin_user_id"):
        raise HTTPException(status_code=401, detail="Chưa đăng nhập admin")
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận file Excel (.xlsx, .xls)")
    
    try:
        # Import openpyxl here to avoid import errors if not installed
        from openpyxl import load_workbook
        
        # Read file content
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), read_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.lower() if cell.value else "")
        
        # Column mapping from XLSX to database fields
        column_map = {
            'title': 'title',
            'authors': 'authors',
            'genres_primary': 'category',
            'pages': 'pages',
            'year': 'publish_year',
            'publisher': 'publisher',
            'price_vnd': 'price_vnd',
            'stock': 'stock',
            'rating_avg': 'rating_avg',
            'short_summary': 'description',
            'image_url': 'image_url'
        }
        
        results = {
            "success": 0,
            "errors": 0,
            "error_details": [],
            "imported_books": []
        }
        
        # Process each row (skip header)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Build data dict from row
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        if header in column_map:
                            row_data[column_map[header]] = value
                
                # Skip empty rows
                if not row_data.get('title'):
                    continue
                
                # Get or create category
                category = None
                if row_data.get('category'):
                    category = get_or_create_category(db, str(row_data['category']))
                
                # Create book
                book = Book(
                    title=str(row_data.get('title', '')),
                    authors=str(row_data.get('authors', 'Unknown')),
                    description=str(row_data.get('description', '')) if row_data.get('description') else None,
                    price_vnd=float(row_data.get('price_vnd', 0)) if row_data.get('price_vnd') else 0,
                    stock=int(row_data.get('stock', 0)) if row_data.get('stock') else 0,
                    image_url=str(row_data.get('image_url', '')) if row_data.get('image_url') else None,
                    rating_avg=float(row_data.get('rating_avg', 0)) if row_data.get('rating_avg') else 0.0,
                    pages=int(row_data.get('pages', 0)) if row_data.get('pages') else None,
                    publisher=str(row_data.get('publisher', '')) if row_data.get('publisher') else None,
                    publish_year=int(row_data.get('publish_year', 0)) if row_data.get('publish_year') else None,
                    is_active=True,
                    category_id=category.id if category else None
                )
                
                db.add(book)
                db.commit()
                db.refresh(book)
                
                # Send webhook
                try:
                    await send_webhook("create", book)
                except Exception as w_e:
                    print(f"Webhook error for import row {row_idx}: {w_e}")

                results["success"] += 1
                results["imported_books"].append({
                    "id": book.id,
                    "title": book.title,
                    "category": category.name if category else None
                })
                
            except Exception as e:
                db.rollback()
                results["errors"] += 1
                results["error_details"].append({
                    "row": row_idx,
                    "error": str(e)
                })
        
        workbook.close()
        
        return JSONResponse(content={
            "message": f"Import hoàn tất: {results['success']} thành công, {results['errors']} lỗi",
            "results": results
        })
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Thư viện openpyxl chưa được cài đặt. Chạy: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý file: {str(e)}")
